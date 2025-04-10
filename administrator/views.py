from .models import Item, ItemType, ItemImage, PartnershipFiles, Partnership, Expenditure, Year, InternationalVisitor, Sector, Center, Office, Purpose
from room_reservation.models import Reservation
from django.core.mail import EmailMessage
from django.views.decorators.http import require_POST
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .forms import ItemForm, ItemTypeForm, ItemImageForm, PartnershipForm, ExpenditureForm
from reportlab.lib.pagesizes import legal, landscape
from django.views.decorators.csrf import csrf_exempt
from room_reservation.models import Room
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from django.core.mail import send_mail, BadHeaderError
from django.utils.html import strip_tags
from smtplib import SMTPException
from django.core.paginator import Paginator
from django.contrib.auth import logout
from decimal import Decimal
from openpyxl.styles import Alignment
from collections import defaultdict
from django.http import HttpResponse, HttpResponseBadRequest
from collections import Counter
from django.db import IntegrityError
from reportlab.lib.units import inch
from django.contrib import messages
from django.http import HttpResponse
from django.http import FileResponse
from reportlab.lib import colors
from django.conf import settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from datetime import datetime
from django.http import JsonResponse
from django.db.models import Count
import pycountry
import shutil
import sqlite3
import os
import json
import logging


def administrator_dashboard(request):
    item_count = Item.objects.filter(is_archived=False).count()
    partnership_count = Partnership.objects.filter(is_removed_from_list=False).count()
    item_type_count = ItemType.objects.count()
    ppmp_per_year_count = Expenditure.objects.values("year").annotate(count=Count("id")).count()

    # Get partnership status distribution
    status_distribution = (
        Partnership.objects
        .filter(is_removed_from_list=False)
        .values("status")
        .annotate(count=Count("id"))
    )

    # Define colors for each status
    status_colors = {
        "Active": "rgba(76, 175, 80, 0.9)",  # Green
        "Pending": "rgba(255, 165, 0, 0.9)",  # Amber
        "Completed": "rgba(33, 150, 243, 0.9)",  # Blue
        "Inactive": "rgba(244, 67, 54, 0.9)"  # Red
    }

    # Extract labels and counts for partnership status
    status_labels = [entry["status"] for entry in status_distribution]
    status_counts = [entry["count"] for entry in status_distribution]
    status_colors_filtered = {label: status_colors.get(label, "rgba(200, 200, 200, 0.7)") for label in status_labels}

    # Category distribution for items
    category_distribution = (
        Item.objects.filter(is_archived=False)
        .values("unit_price")
        .annotate(count=Count("id"))
    )

    # Process category counts
    category_counts = {
        "Property, Plant and Equipment": 0,
        "High Value Semi-Expendable": 0,
        "Low Value Semi-Expendable": 0
    }

    for entry in category_distribution:
        unit_price = entry["unit_price"]
        count = entry["count"]

        if unit_price >= 50000:
            category_counts["Property, Plant and Equipment"] += count
        elif 5000 < unit_price < 50000:
            category_counts["High Value Semi-Expendable"] += count
        else:
            category_counts["Low Value Semi-Expendable"] += count

    # Define colors for each category
    category_colors = {
        "Property, Plant and Equipment": "rgba(33, 150, 243, 0.9)",  # Blue
        "High Value Semi-Expendable": "rgba(255, 165, 0, 0.9)",  # Amber
        "Low Value Semi-Expendable": "rgba(76, 175, 80, 0.9)",  # Green
    }

    # Prepare context for rendering
    context = {
        "item_count": item_count,
        "partnership_count": partnership_count,
        "item_type_count": item_type_count,
        "ppmp_per_year_count": ppmp_per_year_count,
        "status_labels": json.dumps(status_labels),
        "status_counts": json.dumps(status_counts),
        "status_colors_json": json.dumps(status_colors_filtered),
        "category_labels": json.dumps(list(category_counts.keys())),
        "category_counts": json.dumps(list(category_counts.values())),
        "category_colors_json": json.dumps(category_colors),
    }

    return render(request, "administrator/administrator_dashboard.html", context)

def item_list(request):
    if request.method == 'POST':
        form = ItemForm(request.POST)
        image_form = ItemImageForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                item = form.save()
                images = request.FILES.getlist('images')
                for image in images:
                    ItemImage.objects.create(item=item, image=image)
                messages.success(request, 'Item added successfully!')
                return redirect('item_list')
            except IntegrityError:
                messages.error(request, 'An item with this property number already exists.')
        else:
            messages.error(request, 'There was an error adding the item. Please check the form fields.')
    else:
        form = ItemForm()
        image_form = ItemImageForm()

    query = request.GET.get('q', '')
    item_type_id = request.GET.get('item_type', '')

    # Fetch items based on search and filter
    items = Item.objects.filter(is_archived=False).order_by("-id")
    if query:
        items = items.filter(property_number__icontains=query)
    if item_type_id:
        items = items.filter(item_type__id=item_type_id)

    # Get the item type name
    item_type_name = ""
    if item_type_id:
        item_type = ItemType.objects.filter(id=item_type_id).first()
        item_type_name = item_type.name if item_type else ""

    # Pagination
    paginator = Paginator(items, 8)
    page_number = request.GET.get('page')
    page_items = paginator.get_page(page_number)

    item_types = ItemType.objects.all().order_by('name')
    context = {
        'items': page_items,
        'form': form,
        'image_form': image_form,
        'item_types': item_types,
        'query': query,
        'item_type_name': item_type_name,  # Pass the item type name
        'item_type_id': item_type_id,
        'export_url': reverse('export_items_to_pdf') + f'?q={query}&item_type={item_type_id}',
    }
    return render(request, 'administrator/supply_property_inventory.html', context)

def item_type_list(request):
    query = request.GET.get('q', '')  # Get search query
    itemtypes_list = ItemType.objects.all().order_by('name')

    if query:
        itemtypes_list = itemtypes_list.filter(name__icontains=query)  # Filter by search query

    # Pagination
    paginator = Paginator(itemtypes_list, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    itemtypes = paginator.get_page(page_number)

    if request.method == 'POST':
        if 'id' in request.POST:
            # Editing an existing item type
            itemtype = ItemType.objects.get(id=request.POST['id'])
            form = ItemTypeForm(request.POST, request.FILES, instance=itemtype)
            if form.is_valid():
                name = form.cleaned_data.get('name')  # Unique field
                if ItemType.objects.filter(name=name).exclude(id=itemtype.id).exists():
                    messages.error(request, 'An item type with this name already exists.')
                else:
                    form.save()
                    messages.success(request, 'Item type updated successfully!')
                    return redirect('item_type_list')
            else:
                messages.error(request, 'Failed to update item type. Please try again.')
        else:
            # Adding a new item type
            form = ItemTypeForm(request.POST, request.FILES)
            if form.is_valid():
                name = form.cleaned_data.get('name')
                if ItemType.objects.filter(name=name).exists():
                    messages.error(request, 'An item type with this name already exists.')
                else:
                    form.save()
                    messages.success(request, 'Item type added successfully!')
                    return redirect('item_type_list')
            else:
                messages.error(request, 'An error has occured. Please try again later.')
    else:
        form = ItemTypeForm()

    return render(request, 'administrator/item_type_list.html', {
        'itemtypes': itemtypes,
        'form': form,
        'query': query,  # Pass query to template
    })

def edit_item_type(request, pk):
    item_type = get_object_or_404(ItemType, pk=pk)

    if request.method == "POST":
        form = ItemTypeForm(request.POST, request.FILES, instance=item_type)
        if form.is_valid():
            form.save()  # Save the updated item type (name and image)
            messages.success(request, f"Item type '{item_type.name}' updated successfully.")
        else:
            messages.error(request, "Failed to update item type. Please try again.")
        return redirect('item_type_list')  # Redirect to the list view after editing

    form = ItemTypeForm(instance=item_type)  # If GET, populate the form with existing item type details
    return render(request, 'administrator/item_type_list.html', {'form': form, 'item_type': item_type})

def delete_item_type(request, pk):
    item_type = get_object_or_404(ItemType, pk=pk)
    if request.method == "POST":
        messages.get_messages(request).used = True  # Clear previous messages
        try:
            item_type.delete()
            messages.success(request, f"Item type '{item_type.name}' has been deleted successfully.")
        except IntegrityError:
            messages.error(request, f"Cannot delete '{item_type.name}' because there are related records.")
        return redirect('item_type_list')  # Redirect to the item type list page

    return redirect('item_type_list')

def item_detail(request, id):
    item = get_object_or_404(Item, id=id)
    item_types = ItemType.objects.all().order_by("name")  # Fetch all item types
    return render(request, 'administrator/item_detail.html', {'item': item, 'item_types': item_types})

def item_edit(request, id):
    item = get_object_or_404(Item, id=id)

    if request.method == "POST":
        try:
            item.sector_office_division_college = request.POST['sector_office_division_college']
            item.operating_unit_project = request.POST['operating_unit_project']
            item.property_number = request.POST['property_number']
            item.responsible_person = request.POST['responsible_person']
            item.quantity = request.POST['quantity']
            item.unit = request.POST['unit']

            # Fetch the ItemType instance before assigning it
            item_type_id = request.POST['item_type']
            item.item_type = get_object_or_404(ItemType, id=item_type_id)

            item.description = request.POST['description']
            item.date_acquired = request.POST['date_acquired']
            item.fund = request.POST['fund']
            item.ppe_class = request.POST['ppe_class']
            item.estimated_useful_life = request.POST['estimated_useful_life']
            item.unit_price = request.POST['unit_price']
            item.total_amount = request.POST['total_amount']
            item.remarks = request.POST['remarks']  # Add remarks field
            item.is_archived = 'is_archived' in request.POST
            
            item.save()

            # Save uploaded images
            if request.FILES.getlist('images'):
                for image in request.FILES.getlist('images'):
                    ItemImage.objects.create(item=item, image=image)

            messages.success(request, "Item successfully updated.")
            return redirect('item_detail_admin', id=item.id)
        except Exception as e:
            messages.error(request, f"Error updating item: {str(e)}")

    item_types = ItemType.objects.all()
    return render(request, 'administrator/item_detail.html', {'item': item, 'item_types': item_types})

def add_item_image(request, id):
    item = get_object_or_404(Item, id=id)
    
    if request.method == "POST":
        try:
            if request.FILES.getlist('images'):
                for image in request.FILES.getlist('images'):
                    ItemImage.objects.create(item=item, image=image)
            messages.success(request, "Image(s) added successfully.")
        except Exception as e:
            messages.error(request, f"Error adding image(s): {str(e)}")
    
    return redirect('item_edit', id=item.id)

def edit_item_image(request, id):
    item = get_object_or_404(Item, id=id)

    if request.method == "POST":
        image_id = request.POST.get('image_id')
        new_image = request.FILES.get('new_image')
        
        try:
            image_instance = get_object_or_404(ItemImage, id=image_id, item=item)
            if new_image:
                image_instance.image = new_image
                image_instance.save()
                messages.success(request, "Image updated successfully.")
            else:
                messages.error(request, "Please select a new image.")
        except Exception as e:
            messages.error(request, f"Error updating image: {str(e)}")
    
    return redirect('item_detail_admin', id=item.id)

def delete_item_image(request, id):
    item = get_object_or_404(Item, id=id)

    if request.method == "POST":
        image_id = request.POST.get('image_id')
        
        try:
            image_instance = get_object_or_404(ItemImage, id=image_id, item=item)
            image_instance.delete()  # Delete the image instance
            messages.success(request, "Image deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting image: {str(e)}")

    return redirect('item_detail_admin', id=item.id)

def export_items_to_excel(request):
    # Create a new workbook and active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Items'

    # Define headers for the sheet
    headers = [
        "Sector/Office/Division/College", "Operating Unit/Project", "Property Number", 
        "Responsible Person", "Quantity", "Unit", "Type", "Item Description", 
        "Date Acquired", "Fund", "PPE Class", "Estimated Useful Life", 
        "Unit Price (₱)", "Total Amount (₱)"
    ]
    sheet.append(headers)

    # Filters from the GET request
    query = request.GET.get('q', '')
    item_type_id = request.GET.get('item_type', '')

    items = Item.objects.all().order_by("-date_acquired")

    # Apply search filter
    if query:
        items = items.filter(property_number__icontains=query)

    # Apply item type filter and get item type name for filename
    item_type_name = "Items"
    if item_type_id:
        item_type = ItemType.objects.filter(id=item_type_id).first()
        if item_type:
            item_type_name = item_type.name
            items = items.filter(item_type=item_type)

    # Write data to the sheet
    for item in items:
        unit_price = f"₱{item.unit_price:,.2f}" if item.unit_price else ''
        total_amount = f"₱{item.total_amount:,.2f}" if item.total_amount else ''
        
        sheet.append([
            item.sector_office_division_college,
            item.operating_unit_project,
            item.property_number,
            item.responsible_person,
            item.quantity,
            item.unit,
            item.item_type.name if item.item_type else '',
            item.description,
            item.date_acquired.strftime("%B %d, %Y") if item.date_acquired else '',
            item.fund,
            item.ppe_class,
            f"{item.estimated_useful_life} years" if item.estimated_useful_life else '',
            unit_price,
            total_amount
        ])

    # Auto-adjust column widths and wrap text for description
    for col_num, column_cells in enumerate(sheet.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                    if col_letter == get_column_letter(8):  # Wrap text for 'Item Description'
                        cell.alignment = Alignment(wrap_text=True)
            except Exception as e:
                print(e)
        sheet.column_dimensions[col_letter].width = max_length + 2

    # Set the response to send an Excel file with a dynamic filename
    filename = f"{item_type_name.replace(' ', '_')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response

def archive_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    item.is_archived = True
    item.save()
    messages.success(request, f"The item '{item.property_number}' has been archived successfully.")
    return redirect('item_list')  


# Function to wrap text inside a Paragraph for better visibility
def wrap_text(text, font_name='Helvetica'):
    """Wrap text inside a Paragraph to avoid truncating."""
    style = ParagraphStyle(name='Normal', fontSize=7, wordWrap='LTR', fontName=font_name)
    return Paragraph(text, style)

def export_items_to_pdf(request):
    # Get search filters from GET request
    query = request.GET.get('q', '')
    item_type = request.GET.get('item_type', '')

    # Create HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="items_report.pdf"'
    
    # Create a PDF document with reduced margins and landscape legal size
    doc = SimpleDocTemplate(response, pagesize=landscape(legal),
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch, 
                            topMargin=0.5 * inch, bottomMargin=0.3 * inch)
    
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    header_style = styles['BodyText']
    header_style.fontSize = 7
    header_style.leading = 9
    header_style.fontName = 'Helvetica-Bold'  # Use Helvetica-Bold for header
    header_style.textColor = colors.whitesmoke  # Set text color to white

    # Shortened headers with wrapping
    def header_text(text):
        return Paragraph(text, header_style)
    
    # Prepare the header data
    data = [
        [
            header_text("Sector/<br/>Office/Division"),
            header_text("Operating <br/>Unit/Project"),
            "Prop. No.",
            "Resp. Person",
            "Qty", "Unit", "Type", 
            header_text("Item Description"),
            "Date Acquired", "Fund", "PPE Class", 
            "Useful Life", "Unit Price", "Total"
        ]
    ]
    
    # Fetch items and apply search filters
    items = Item.objects.all()

    # Apply search filter
    if query:
        items = items.filter(property_number__icontains=query)

    # Apply item type filter
    if item_type:
        items = items.filter(item_type__id=item_type)

    # Wrap each field for each item
    for item in items:
        data.append([ 
            wrap_text(item.sector_office_division_college),
            wrap_text(item.operating_unit_project),
            wrap_text(item.property_number),
            wrap_text(item.responsible_person),
            item.quantity, item.unit,
            wrap_text(item.item_type.name),
            wrap_text(item.description),
            item.date_acquired.strftime('%B %d, %Y'),
            wrap_text(item.fund),
            wrap_text(item.ppe_class),
            f"{item.estimated_useful_life} yrs",
            wrap_text(f"P{item.unit_price:,.2f}"),  # Removed custom font for unit price
            wrap_text(f"P{item.total_amount:,.2f}")  # Removed custom font for total
        ])
    
    # Adjust column widths
    col_widths = [
        0.9 * inch, 0.9 * inch, 1.2 * inch,
        1.2 * inch, 0.4 * inch, 0.5 * inch, 
        0.9 * inch, 1.8 * inch, 1.0 * inch,
        0.9 * inch, 0.9 * inch, 0.7 * inch,
        0.8 * inch, 0.8 * inch
    ]
    
    # Create and style the table
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([ 
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Keep using Helvetica-Bold
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


def archived_items_view(request):
    # Filter archived items
    archived_items = Item.objects.filter(is_archived=True)

    # Paginate the results
    from django.core.paginator import Paginator
    paginator = Paginator(archived_items, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the items to the template
    return render(request, 'administrator/archived_items.html', {'items': page_obj})

def unarchive_item(request, item_id):
    # Ensure that the user has the proper permissions
    item = get_object_or_404(Item, id=item_id)
    
    # Check if the item is archived before allowing unarchiving
    if not item.is_archived:
        # Item is not archived, show an error message
        messages.error(request, "This item is not archived and cannot be restored.")
        return redirect('item_list')  # Change to the appropriate view if needed
    
    # Unarchive the item
    item.is_archived = False
    item.save()

    # Show a success message
    messages.success(request, f"The item '{item.property_number}' has been successfully restored.")

    # Redirect back to the item list or relevant page
    return redirect('item_list')  # Change 'item_list' to your relevant view


@login_required    
def download_db_backup(request):
    db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
    if os.path.exists(db_path):
        # Format the date in 'YYYY-MM-DD' format
        date_str = datetime.now().strftime("%B %d, %Y")
        filename = f"db_backup_{date_str}.sqlite3"
        
        response = FileResponse(open(db_path, 'rb'), content_type='application/x-sqlite3')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    else:
        return HttpResponse("Database file not found.", status=404)

def restore_database(request):
    if request.method == 'POST' and request.FILES.get('db_file'):
        db_file = request.FILES['db_file']
        
        # Confirm file type is correct
        if not db_file.name.endswith('.sqlite3'):
            messages.error(request, "Please upload a valid SQLite3 database file (.sqlite3).")
            return redirect('restore_database')

        # Paths for backup and current database
        backup_path = os.path.join(settings.BASE_DIR, 'db_backup.sqlite3')
        current_db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
        temp_db_path = os.path.join(settings.BASE_DIR, 'temp_uploaded_db.sqlite3')

        try:
            # Save uploaded file temporarily
            with open(temp_db_path, 'wb+') as temp_db:
                for chunk in db_file.chunks():
                    temp_db.write(chunk)
            
            # Check table structure in current and uploaded databases
            def get_table_names(db_path):
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                tables = {table[0] for table in cursor.fetchall()}
                conn.close()
                return tables

            current_tables = get_table_names(current_db_path)
            uploaded_tables = get_table_names(temp_db_path)
            
            # Verify table structure
            if current_tables != uploaded_tables:
                messages.error(request, "The uploaded database does not have the same structure as the current database.")
                os.remove(temp_db_path)  # Remove temporary uploaded database file
                return redirect('restore_database')

            # Backup current database
            if os.path.exists(current_db_path):
                shutil.copy(current_db_path, backup_path)
            
            # Replace current database with uploaded file
            shutil.move(temp_db_path, current_db_path)
            messages.success(request, "Database restored successfully!")

        except Exception as e:
            # Restore from backup if something goes wrong
            if os.path.exists(backup_path):
                shutil.copy(backup_path, current_db_path)

            messages.error(request, f"An error occurred while restoring: {str(e)}")
            return redirect('restore_database')

        return redirect('restore_database')

    return render(request, 'administrator/restore_database.html')

def logout_view(request):
    logout(request)
    messages.success(request, "You have been Logged-Out Successfully!")
    return redirect('login_url')


def partnership_list(request):
    if request.method == 'POST':
        try:
            # Get the data from the POST request
            partner = request.POST.get('partner')
            url = request.POST.get('url')
            country = request.POST.get('country')
            continent = request.POST.get('continent')
            type_of_organization = request.POST.get('type_of_organization')
            description = request.POST.get('description')
            status = request.POST.get('status')
            logo = request.FILES.get('logo')  # Handle the logo file upload
            university_files = request.FILES.getlist('university_files')  # Handle multiple files

            # Create the partnership object
            partnership = Partnership.objects.create(
                partner=partner,
                url=url,
                country=country,
                continent=continent,
                type_of_organization=type_of_organization,
                description=description,
                status=status,
                logo=logo
            )

            # Save each file as a PartnershipFiles object
            for file in university_files:
                PartnershipFiles.objects.create(
                    partnership=partnership,
                    file=file
                )

            messages.success(request, 'Partnership added successfully!')
        except Exception as e:
            messages.error(request, f'Failed to add partnership: {str(e)}')

        return redirect('partnership_list')

    # Handling GET request (displaying the partnership list)
    query = request.GET.get('q', '')
    continent_filter = request.GET.get('continent', '')
    status_filter = request.GET.get('status', '')  # Get the status filter from the request

    partnerships = Partnership.objects.all().order_by("partner").filter(is_removed_from_list=False)

    # Apply search filter
    if query:
        partnerships = partnerships.filter(
            partner__icontains=query
        ) | partnerships.filter(
            country__icontains=query
        ) | partnerships.filter(
            type_of_organization__icontains=query
        )

    # Apply continent filter
    if continent_filter:
        partnerships = partnerships.filter(continent=continent_filter)

    # Apply status filter
    if status_filter:
        partnerships = partnerships.filter(status=status_filter)

    # Apply pagination
    paginator = Paginator(partnerships, 10)  # Show 10 partnerships per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'partnerships': page_obj,
        'query': query,
        'continent_filter': continent_filter,
        'status_filter': status_filter,  # Pass the status filter to the template
        'CONTINENT_CHOICES': Partnership.CONTINENT_CHOICES,
        'STATUS_CHOICES': Partnership.STATUS_CHOICES,  # Pass status choices to the template
    }

    return render(request, 'administrator/partnership_list.html', context)

def partnership_detail(request, partnership_id):
    partnership = get_object_or_404(Partnership, id=partnership_id)
    partnership_files = partnership.files.all()  # Get the files related to this partnership

    # Pass choices for continent and status to the template
    continent_choices = Partnership.CONTINENT_CHOICES
    status_choices = Partnership.STATUS_CHOICES

    if request.method == 'POST':
        partnership.partner = request.POST.get('partner')
        partnership.country = request.POST.get('country')
        partnership.continent = request.POST.get('continent')
        partnership.type_of_organization = request.POST.get('type_of_organization')
        partnership.description = request.POST.get('description')
        partnership.status = request.POST.get('status')
        
        if request.FILES.get('logo'):
            partnership.logo = request.FILES['logo']
        
        partnership.save()
        messages.success(request, 'Partnership updated successfully!')
        return redirect('partnership_detail', partnership_id=partnership.id)

    context = {
        'partnership': partnership,
        'partnership_files': partnership_files,
        'continent_choices': continent_choices,
        'status_choices': status_choices,
    }
    return render(request, 'administrator/partnership_detail.html', context)


def upload_file_to_partnership(request, partnership_id):
    if request.method == 'POST':
        partnership = get_object_or_404(Partnership, id=partnership_id)
        file = request.FILES.get('file')
        description = request.POST.get('description')

        if file:
            PartnershipFiles.objects.create(
                partnership=partnership,
                file=file,
                description=description
            )
            messages.success(request, "File uploaded successfully!")
        else:
            messages.error(request, "File is required.")

    return redirect('partnership_list')  # Redirecting to partnership list


@csrf_exempt
def delete_partnership_file(request):
    if request.method == 'POST':
        file_id = request.POST.get('file_id')
        if not file_id:
            return JsonResponse({'success': False, 'error': 'Missing file_id'})
        
        try:
            file = PartnershipFiles.objects.get(id=file_id)
            file.delete()  # Delete the file from the database and storage

          
            # Send a success response back to the client
            return JsonResponse({'success': True})
        
        except PartnershipFiles.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'File not found'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def expenditure_list(request):
    # Group expenditures by classification
    expenditures = Expenditure.objects.all().order_by("classification", "expenditure_type", "item_number")
    grouped_expenditures = {}
    
    for expenditure in expenditures:
        classification = expenditure.classification
        if classification not in grouped_expenditures:
            grouped_expenditures[classification] = {}
        
        expenditure_type = expenditure.expenditure_type
        if expenditure_type not in grouped_expenditures[classification]:
            grouped_expenditures[classification][expenditure_type] = []

        grouped_expenditures[classification][expenditure_type].append(expenditure)

    # Pass the grouped expenditures and classification choices to the template
    context = {
        "grouped_expenditures": grouped_expenditures,
        "classification_choices": Expenditure.CLASSIFICATION_CHOICES,  # Add classification choices
    }
    return render(request, "administrator/expenditure_details.html", context)


def add_expenditure(request, year_id):
    year = get_object_or_404(Year, id=year_id)  # Get the selected year

    if request.method == 'POST':
        # Extract data from the form
        classification = request.POST.get('classification')
        expenditure_type = request.POST.get('expenditure_type')
        description = request.POST.get('description')
        mode_of_procurement = request.POST.get('mode_of_procurement')
        quarter1 = int(request.POST.get('quarter1', 0))
        quarter2 = int(request.POST.get('quarter2', 0))
        quarter3 = int(request.POST.get('quarter3', 0))
        quarter4 = int(request.POST.get('quarter4', 0))
        unit_of_measure = request.POST.get('unit_of_measure')
        unit_price = float(request.POST.get('unit_price'))
        remarks = request.POST.get('remarks', '')

        # Create and save the new expenditure
        try:
            expenditure = Expenditure(
                year=year,  # Associate the expenditure with the selected year
                classification=classification,
                expenditure_type=expenditure_type,
                description=description,
                mode_of_procurement=mode_of_procurement,
                quarter1=quarter1,
                quarter2=quarter2,
                quarter3=quarter3,
                quarter4=quarter4,
                unit_of_measure=unit_of_measure,
                unit_price=unit_price,
                remarks=remarks
            )
            expenditure.save()
            messages.success(request, 'Expenditure added successfully!')
        except Exception as e:
            messages.error(request, f'Error adding expenditure: {str(e)}')

        return redirect('year_detail', year_id=year_id)  # Redirect back to the year detail page

    return redirect('year_detail', year_id=year_id)  # Redirect if not a POST request
def year_list(request):
    years = Year.objects.all().order_by("year")
    current_year = datetime.now().year
    available_years = list(range(current_year - 50, current_year + 51))

    # Prepare expenditure data for each year
    year_data = []
    for year in years:
        expenditures = Expenditure.objects.filter(year=year)
        total_spent = sum(exp.total_amount for exp in expenditures)
        unspent = max(year.budget - total_spent, 0) if year.budget else 0
        
        # Group by classification
        classification_data = {}
        for exp in expenditures:
            classification_data[exp.classification] = classification_data.get(exp.classification, 0) + float(exp.total_amount)
        
        year_data.append({
            'year_obj': year,
            'total_spent': total_spent,
            'unspent': unspent,
            'classification_data': classification_data
        })

    context = {
        'year_data': year_data,
        'available_years': available_years
    }
    return render(request, 'administrator/year_list.html', context)

def add_year(request):
    if request.method == "POST":
        year = request.POST.get("year")
        budget = request.POST.get("budget")
        description = request.POST.get("description", "")

        # Ensure the year is unique
        if Year.objects.filter(year=year).exists():
            messages.error(request, "This year already exists.")
            return redirect("year_list")

        # Create and save the new year
        Year.objects.create(year=year, budget=budget, description=description)
        messages.success(request, "New year added successfully!")
        return redirect("year_list")

    return redirect("year_list")

def year_detail(request, year_id):
    year = get_object_or_404(Year, id=year_id)
    expenditures = Expenditure.objects.filter(year=year).order_by("classification", "expenditure_type", "item_number")

    grouped_expenditures = {}
    subtotals = defaultdict(lambda: defaultdict(lambda: Decimal(0)))
    totals_per_classification = defaultdict(Decimal)
    grand_total = Decimal(0)

    for expenditure in expenditures:
        classification = expenditure.classification
        expenditure_type = expenditure.expenditure_type

        if classification not in grouped_expenditures:
            grouped_expenditures[classification] = {}

        if expenditure_type not in grouped_expenditures[classification]:
            grouped_expenditures[classification][expenditure_type] = []

        grouped_expenditures[classification][expenditure_type].append(expenditure)

        # Compute subtotal for expenditure type
        subtotals[classification][expenditure_type] += Decimal(expenditure.total_amount)

        # Compute total per classification
        totals_per_classification[classification] += Decimal(expenditure.total_amount)

        # Compute grand total
        grand_total += Decimal(expenditure.total_amount)

    context = {
    "grouped_expenditures": grouped_expenditures,
    "subtotals": {k: dict(v) for k, v in subtotals.items()},  # Convert nested defaultdict to dict
    "totals_per_classification": dict(totals_per_classification),  # Convert defaultdict to dict
    "grand_total": grand_total,
    "classification_choices": Expenditure.CLASSIFICATION_CHOICES,
    "selected_year": year,
    }

    return render(request, "administrator/year_detail.html", context)

def delete_expenditure(request, expenditure_id):
    try:
        expenditure = get_object_or_404(Expenditure, id=expenditure_id)
        year_id = expenditure.year.id  # Get the year ID before deleting the expenditure
        
        expenditure.delete()
        messages.success(request, "Expenditure deleted successfully.")

    except Expenditure.DoesNotExist:
        messages.error(request, "Expenditure not found.")
    
    except IntegrityError:
        messages.error(request, "Cannot delete this expenditure due to database integrity constraints.")

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")

    return redirect("year_detail", year_id=year_id)  # Redirect to year_detail

def remove_partnership(request, partnership_id):
    partnership = get_object_or_404(Partnership, id=partnership_id)
    try:
        partnership.is_removed_from_list = True
        partnership.save()
        messages.success(request, 'Partnership has been successfully removed from the list.')
    except Exception as e:
        messages.error(request, f'An error occurred while removing the partnership: {str(e)}')
    
    return redirect(reverse('partnership_list'))

def removed_partnerships_view(request):
    # Filter partnerships that are removed from the list
    removed_partnerships = Partnership.objects.filter(is_removed_from_list=True)

    return render(request, 'administrator/removed_partnerships.html', {
        'removed_partnerships': removed_partnerships,
    })

def edit_partnership(request, partnership_id):
    partnership = get_object_or_404(Partnership, id=partnership_id)

    if request.method == 'POST':
        form = PartnershipForm(request.POST, request.FILES, instance=partnership)
        if form.is_valid():
            form.save()
            return redirect(request.META.get('HTTP_REFERER', 'home')) 

    return redirect(request.META.get('HTTP_REFERER', 'home'))  

def visitor_list(request):
    # Fetch all visitors
    visitors_list = InternationalVisitor.objects.all().order_by('-id')

    # Search by name
    search_query = request.GET.get('search', '').strip()
    if search_query:
        visitors_list = visitors_list.filter(name__icontains=search_query)

    # Filter by country (only countries present in the database)
    country_filter = request.GET.get('country', '')
    if country_filter:
        visitors_list = visitors_list.filter(country=country_filter)

    # Filter by year
    year_filter = request.GET.get('year', '')
    if year_filter:
        visitors_list = visitors_list.filter(year=year_filter)

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(visitors_list, 10)  # Show 10 visitors per page
    visitors = paginator.get_page(page)

    # Count sector visits
    sector_counter = Counter()
    for visitor in visitors_list:  # Use full visitor list for chart data
        for sector in visitor.sector_visited.all():
            sector_counter[sector.name] += 1

    # Convert sector data for the chart
    sector_labels = list(sector_counter.keys())
    sector_counts = list(sector_counter.values())

    # Define choices (Do not change the original choices)
    countries = sorted(pycountry.countries, key=lambda x: x.name)
    COUNTRY_CHOICES = [(country.name, country.name) for country in countries]
    SECTOR_CHOICES = Sector.objects.all()
    CENTER_CHOICES = Center.objects.all()
    OFFICE_CHOICES = Office.objects.all()
    PURPOSE_CHOICES = Purpose.objects.all()
    YEAR_CHOICES = [(year, str(year)) for year in range(2000, 2031)]

    # Get unique countries from database for filtering
    existing_countries = InternationalVisitor.objects.values_list('country', flat=True).distinct()

    # Get unique years from database for filtering
    existing_years = InternationalVisitor.objects.values_list('year', flat=True).distinct()

    # Context
    context = {
        'visitors': visitors,  # Paginated visitors
        'COUNTRY_CHOICES': COUNTRY_CHOICES,  # Unchanged for editing
        'SECTOR_CHOICES': SECTOR_CHOICES,
        'CENTER_CHOICES': CENTER_CHOICES,
        'OFFICE_CHOICES': OFFICE_CHOICES,
        'PURPOSE_CHOICES': PURPOSE_CHOICES,
        'YEAR_CHOICES': YEAR_CHOICES,  # Unchanged for editing
        'sector_labels': json.dumps(sector_labels),
        'sector_counts': json.dumps(sector_counts),

        # New search and filter context
        'existing_countries': existing_countries,  # Countries present in DB
        'existing_years': existing_years,  # Years present in DB
        'search_query': search_query,
        'country_filter': country_filter,
        'year_filter': year_filter,
    }

    return render(request, 'administrator/list_of_visitors.html', context)

def add_visitor(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        country = request.POST.get('country')
        year = request.POST.get('year')
        sector_visited_ids = request.POST.getlist('sector_visited')  
        center_unit_ids = request.POST.getlist('center_unit')  
        office_college_ids = request.POST.getlist('office_college')  
        purpose_text = request.POST.get('purpose')  # Get purpose as text

        # Store previously entered data in session
        request.session['form_data'] = {
            "entered_name": name,
            "entered_country": country,
            "entered_year": year,
            "selected_sectors": sector_visited_ids,
            "selected_centers": center_unit_ids,
            "selected_offices": office_college_ids,
            "entered_purpose": purpose_text,  # Store purpose text
        }

        if not country or not year:
            messages.error(request, "Country and Year are required!")
            return redirect('add_visitor')  # Reload the form to show the error

        try:
            # Create the visitor
            visitor = InternationalVisitor.objects.create(
                name=name,
                country=country,
                year=year,
                purpose=purpose_text  # Save purpose as text
            )

            # Add many-to-many relationships
            visitor.sector_visited.set(Sector.objects.filter(id__in=sector_visited_ids))
            visitor.center_unit.set(Center.objects.filter(id__in=center_unit_ids))
            visitor.office_college.set(Office.objects.filter(id__in=office_college_ids))

            messages.success(request, f"Visitor '{name}' from {country} added successfully!")
            return redirect('add_visitor')  # Stay on the same page, keeping input
        except Exception as e:
            messages.error(request, f"Failed to add visitor '{name}' from {country}: {str(e)}")
            return redirect('add_visitor')  # Stay on the form in case of failure

    # Retrieve previously entered data from session
    form_data = request.session.pop('form_data', {})  # Get data and remove from session

    # Fetch all visitors (with search and filtering capabilities)
    visitors_list = InternationalVisitor.objects.all().order_by('-id')

    # Search by name
    search_query = request.GET.get('search', '').strip()
    if search_query:
        visitors_list = visitors_list.filter(name__icontains=search_query)

    # Filter by country (only countries present in the database)
    country_filter = request.GET.get('country', '')
    if country_filter:
        visitors_list = visitors_list.filter(country=country_filter)

    # Filter by year
    year_filter = request.GET.get('year', '')
    if year_filter:
        visitors_list = visitors_list.filter(year=year_filter)

    # Pagination for visitor list
    page = request.GET.get('page', 1)
    paginator = Paginator(visitors_list, 10)  # Show 10 visitors per page
    visitors = paginator.get_page(page)

    # Define choices (Unchanged for filtering and form inputs)
    countries = sorted(pycountry.countries, key=lambda x: x.name)
    COUNTRY_CHOICES = [(country.name, country.name) for country in countries]
    SECTOR_CHOICES = Sector.objects.all()
    CENTER_CHOICES = Center.objects.all()
    OFFICE_CHOICES = Office.objects.all()
    YEAR_CHOICES = [(year, str(year)) for year in range(2000, 2031)]

    # Get unique countries and years from database for filtering options
    existing_countries = InternationalVisitor.objects.values_list('country', flat=True).distinct()
    existing_years = InternationalVisitor.objects.values_list('year', flat=True).distinct()

    # Context includes form values, visitors list, and filtering options
    context = {
        "entered_name": form_data.get("entered_name", ""),
        "entered_country": form_data.get("entered_country", ""),
        "entered_year": form_data.get("entered_year", ""),
        "selected_sectors": form_data.get("selected_sectors", []),
        "selected_centers": form_data.get("selected_centers", []),
        "selected_offices": form_data.get("selected_offices", []),
        "entered_purpose": form_data.get("entered_purpose", ""),  # Pass purpose text

        # Choices for form inputs and filtering
        "COUNTRY_CHOICES": COUNTRY_CHOICES,
        "SECTOR_CHOICES": SECTOR_CHOICES,
        "CENTER_CHOICES": CENTER_CHOICES,
        "OFFICE_CHOICES": OFFICE_CHOICES,
        "YEAR_CHOICES": YEAR_CHOICES,

        # Visitors and filters
        "visitors": visitors,
        "existing_countries": existing_countries,
        "existing_years": existing_years,
        "search_query": search_query,
        "country_filter": country_filter,
        "year_filter": year_filter,
    }

    return render(request, "administrator/list_of_visitors.html", context)



def edit_visitor(request):
    if request.method == 'POST':
        visitor_id = request.POST.get('visitor_id')
        visitor = get_object_or_404(InternationalVisitor, id=visitor_id)
        
        # Get the current page number from the form
        current_page = request.POST.get('current_page', '1')
        
        # Track changes
        changes = [f"Visitor: {visitor.name}"]  # Always include the visitor's name

        # Update basic fields and check for changes
        name = request.POST.get('name')
        country = request.POST.get('country')
        year = request.POST.get('year')
        purpose = request.POST.get('purpose')

        if visitor.name != name:
            changes.append(f"Name changed from {visitor.name} to {name}")
            visitor.name = name
        
        if visitor.country != country:
            changes.append(f"Country changed from {visitor.country} to {country}")
            visitor.country = country
        
        if visitor.year != year:
            changes.append(f"Year changed from {visitor.year} to {year}")
            visitor.year = year
        
        if visitor.purpose != purpose:
            changes.append(f"Purpose changed from '{visitor.purpose}' to '{purpose}'")
            visitor.purpose = purpose

        visitor.save()

        # Update many-to-many fields
        sector_visited = request.POST.getlist('sector_visited')
        center_unit = request.POST.getlist('center_unit')
        office_college = request.POST.getlist('office_college')

        if set(visitor.sector_visited.values_list('id', flat=True)) != set(map(int, sector_visited)):
            changes.append("Sectors Visited updated")
            visitor.sector_visited.set(sector_visited)

        if set(visitor.center_unit.values_list('id', flat=True)) != set(map(int, center_unit)):
            changes.append("Centers/Units updated")
            visitor.center_unit.set(center_unit)

        if set(visitor.office_college.values_list('id', flat=True)) != set(map(int, office_college)):
            changes.append("Offices/Colleges updated")
            visitor.office_college.set(office_college)

        messages.success(request, "Visitor updated successfully!\n" + "\n".join(changes))

        # Redirect back to the same page with the page parameter
        return redirect(f"{reverse('visitor_list')}?page={current_page}")
    else:
        messages.error(request, 'Invalid request method.')
        return redirect('visitor_list')
    
def delete_visitor(request, visitor_id):
    visitor = get_object_or_404(InternationalVisitor, id=visitor_id)
    visitor_name = visitor.name
    visitor.delete()
    
    messages.success(request, f"Visitor '{visitor_name}' was deleted successfully.")
    
    return JsonResponse({"message": f"Visitor '{visitor_name}' was deleted successfully."})
def export_international_visitors_to_excel(request):
    # Create a new workbook and active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'International Visitors'

    # Define headers
    headers = [
        "ID", "Name", "Country", "Year", "Sector Visited", "Center/Unit", 
        "Office/College", "Purpose"
    ]
    sheet.append(headers)

    # Get filters from the request
    query = request.GET.get('q', '')
    year = request.GET.get('year', '')

    # Retrieve visitors with optional filters
    visitors = InternationalVisitor.objects.all().order_by("-year")
    
    if query:
        visitors = visitors.filter(name__icontains=query)
    if year:
        visitors = visitors.filter(year=year)

    # Write data to the sheet
    for visitor in visitors:
        # Join sectors with a comma and space
        sectors = ", ".join([sector.name for sector in visitor.sector_visited.all()]) or "N/A"
        
        # Join centers and offices with a newline character for better readability
        centers = "\n".join([center.name for center in visitor.center_unit.all()]) or "N/A"
        offices = "\n".join([office.name for office in visitor.office_college.all()]) or "N/A"
        
        # Get purpose directly from the text field
        purpose = visitor.purpose or "N/A"

        sheet.append([
            visitor.id,
            visitor.name,
            visitor.country,
            visitor.year,
            sectors,
            centers,
            offices,
            purpose,  # Now using the direct text field value
        ])

    # Auto-adjust column widths and row heights for better readability
    for col_num, column_cells in enumerate(sheet.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:
                    # Calculate the maximum length of the cell content
                    max_length = max(max_length, len(str(cell.value)))
                    # Enable text wrapping for the cell
                    cell.alignment = Alignment(wrap_text=True)
            except Exception as e:
                print(e)
        # Set column width based on the maximum length of the content
        sheet.column_dimensions[col_letter].width = max_length + 2

    # Adjust row heights only for rows with cells containing multiple lines
    for row in sheet.iter_rows():
        max_lines = 1  # Default to 1 line
        for cell in row:
            if cell.value and "\n" in str(cell.value):
                # Count the number of lines in the cell
                num_lines = len(str(cell.value).split("\n"))
                # Update max_lines if this cell has more lines
                max_lines = max(max_lines, num_lines)
        # Adjust row height only if max_lines > 1
        if max_lines > 1:
            sheet.row_dimensions[row[0].row].height = 15 * max_lines

    # Set response headers for file download
    filename = f"International_Visitors_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)

    return response

def list_of_rooms(request):
    rooms = Room.objects.prefetch_related("reservations").all()
    return render(request, "administrator/reservation/list_of_rooms.html", {"rooms": rooms})

def add_room(request):
    if request.method == "POST":
        room_number = request.POST.get("room_number")
        room_type = request.POST.get("room_type")
        capacity = request.POST.get("capacity")
        description = request.POST.get("description", "")
        is_available = request.POST.get("is_available") == "on"
        image = request.FILES.get("image")

        # Ensure room number is unique
        if Room.objects.filter(room_number=room_number).exists():
            messages.error(request, "Room number already exists.")
            return redirect("list_of_rooms")  # Change this to the actual room list URL

        # Create Room
        Room.objects.create(
            room_number=room_number,
            room_type=room_type,
            capacity=capacity,
            description=description,
            is_available=is_available,
            image=image,
        )

        messages.success(request, "Room added successfully!")
        return redirect("list_of_rooms")  

    return render(request, "administrator/reservations/list_of_rooms.html") 

def delete_room(request, room_id):
    room = get_object_or_404(Room, id=room_id)
    
    try:
        room.delete()
        messages.success(request, f"Room {room.room_number} was successfully deleted.")
    except Exception as e:
        messages.error(request, f"Failed to delete room {room.room_number}. Error: {str(e)}")

    return redirect("list_of_rooms")


def edit_room(request, room_id):
    room = get_object_or_404(Room, id=room_id)

    if request.method == "POST":
        room.room_type = request.POST.get("room_type")
        room.room_number = request.POST.get("room_number")
        room.is_available = request.POST.get("is_available") == "True"
        room.capacity = request.POST.get("capacity")
        room.description = request.POST.get("description")

        if "room_image" in request.FILES:
            room.image = request.FILES["room_image"]

        room.save()
        messages.success(request, f"Room {room.room_number} updated successfully!")
        return redirect("list_of_rooms")  

    return render(request, "administrator/reservations/list_of_rooms.html", {"room": room})

def delete_year(request, year_id):
    year = get_object_or_404(Year, id=year_id)
    year.delete()
    messages.success(request, f"The year {year.year} has been successfully deleted.")
    return redirect(reverse('year_list'))  # Update with the correct URL name for your list view

def edit_year(request):
    if request.method == "POST":
        year_id = request.POST.get("year_id")
        year = get_object_or_404(Year, id=year_id)
        year.year = request.POST.get("year")
        year.description = request.POST.get("description")
        year.budget = request.POST.get("budget")
        year.save()
        messages.success(request, "Year updated successfully!")
        return redirect("year_list")

def edit_expenditure(request, expenditure_id):
    expenditure = get_object_or_404(Expenditure, id=expenditure_id)
    year_id = expenditure.year.id
    
    if request.method == 'POST':
        try:
            # Update all fields from the form
            expenditure.classification = request.POST.get('classification')
            expenditure.expenditure_type = request.POST.get('expenditure_type')
            expenditure.description = request.POST.get('description')
            expenditure.mode_of_procurement = request.POST.get('mode_of_procurement')
            
            # Update quarter values
            expenditure.quarter1 = int(request.POST.get('quarter1', 0) or 0)
            expenditure.quarter2 = int(request.POST.get('quarter2', 0) or 0)
            expenditure.quarter3 = int(request.POST.get('quarter3', 0) or 0)
            expenditure.quarter4 = int(request.POST.get('quarter4', 0) or 0)
            
            # Update unit of measure and price
            expenditure.unit_of_measure = request.POST.get('unit_of_measure')
            expenditure.unit_price = Decimal(request.POST.get('unit_price', 0) or 0)
            
            # Let the model's save() method handle the calculations
            expenditure.remarks = request.POST.get('remarks', '')

            expenditure.save()
            messages.success(request, 'Expenditure updated successfully!')
        except Exception as e:
            messages.error(request, f'Error updating expenditure: {str(e)}')
        
        return redirect('year_detail', year_id=year_id)

    return redirect('year_detail', year_id=year_id)

logger = logging.getLogger(__name__)

@require_POST
def approve_reservation(request):
    reservation_id = request.POST.get('reservation_id')
    confirmation_message = request.POST.get('confirmation_message', '')
    
    try:
        reservation = get_object_or_404(Reservation, id=reservation_id)
        
        if reservation.status != 'Pending':
            messages.error(request, 'This reservation has already been processed.')
            logger.warning(f'Attempt to reprocess reservation {reservation_id}')
            return redirect('pending_reservations')
        
        # Update status and confirmed_at field
        reservation.status = 'Confirmed'
        reservation.confirmed_at = timezone.now()  # Set the confirmed_at datetime field
        
        if confirmation_message:
            reservation.details = confirmation_message
        
        reservation.save()
        
        # Send confirmation email
        email_sent, error_message = send_confirmation_email(reservation, confirmation_message)
        
        if email_sent:
            messages.success(request, f'Reservation approved! Confirmation sent to {reservation.email}.')
            logger.info(f'Successfully approved reservation {reservation_id}')
        else:
            messages.warning(request, f'Reservation approved but email failed: {error_message}')
            logger.error(f'Email failed for reservation {reservation_id}: {error_message}')
        
        return redirect('pending_reservations')
    
    except Exception as e:
        messages.error(request, f'System error: {str(e)}')
        logger.exception(f'System error processing reservation {reservation_id}')
        return redirect('pending_reservations')
    
def send_confirmation_email(reservation, additional_message):
    subject = f"Booking Confirmation - {reservation.room} (Ref: RES-{reservation.id:06d})"
    
    context = {
        'reservation': reservation,
        'additional_message': additional_message,
        'booking_reference': f"RES-{reservation.id:06d}",
        'hotel_contact': settings.DEFAULT_FROM_EMAIL,
    }
    
    try:
        html_message = render_to_string('emails/booking_confirmation.html', context)
        plain_message = strip_tags(html_message)
        
        send_mail(
            subject,
            plain_message,
            settings.DEFAULT_FROM_EMAIL,
            [reservation.email],
            html_message=html_message,
            fail_silently=False,
        )
        return True, ""
    
    except BadHeaderError:
        error_msg = "Invalid email headers"
        logger.error(f'Email header error for reservation {reservation.id}')
        return False, error_msg
    except SMTPException as e:
        error_msg = f"Email server error: {str(e)}"
        logger.error(f'SMTP error for reservation {reservation.id}: {str(e)}')
        return False, error_msg
    except Exception as e:
        error_msg = f"Unexpected error: {str(e)}"
        logger.exception(f'Email error for reservation {reservation.id}')
        return False, error_msg

def confirmed_reservations(request):
    # Get all confirmed reservations ordered by arrival date
    confirmed_reservations_list = Reservation.objects.filter(status='Confirmed').order_by('arrival_date')

    # Convert 'created_at' and 'confirmed_at' to local time for each reservation
    for reservation in confirmed_reservations_list:
        reservation.created_at = timezone.localtime(reservation.created_at)
        reservation.confirmed_at = timezone.localtime(reservation.confirmed_at)

    # Pagination
    paginator = Paginator(confirmed_reservations_list, 10)  # Show 10 reservations per page
    page_number = request.GET.get('page')
    confirmed_reservations = paginator.get_page(page_number)
    
    context = {
        'confirmed_reservations': confirmed_reservations,
    }
    return render(request, 'administrator/reservation/confirmed_reservations.html', context)

def cancel_reservation(request):
    if request.method == 'POST':
        reservation_id = request.POST.get('reservation_id')
        reason = request.POST.get('reason_for_cancelling')
        cancelled_by = request.POST.get('cancelled_by')
        
        try:
            reservation = Reservation.objects.get(id=reservation_id)
            reservation.status = 'Cancelled'
            reservation.reason_for_cancelling = reason
            reservation.cancelled_by = cancelled_by
            reservation.cancelled_at = timezone.now()  # Update the cancelled_at datetime field
            reservation.save()
            
            # Send email notification
            send_cancellation_email(reservation, reason)
            
            messages.success(request, f'Reservation #{reservation_id} has been cancelled successfully. An email has been sent to the guest.')
        except Reservation.DoesNotExist:
            messages.error(request, 'Reservation not found.')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
    
    # Redirect back to the appropriate page
    if request.META.get('HTTP_REFERER'):
        return redirect(request.META['HTTP_REFERER'])
    return redirect('pending_reservations')  # Fallback if no referer

def send_cancellation_email(reservation, reason):
    subject = f"Reservation Cancellation Notification (Ref: #{reservation.id})"
    
    # Context for the email template
    context = {
        'reservation': reservation,
        'reason': reason,
        'contact_email': settings.DEFAULT_FROM_EMAIL,

    }
    
    # Render HTML content
    html_message = render_to_string('emails/reservation_cancelled.html', context)
    plain_message = strip_tags(html_message)
    
    email = EmailMessage(
        subject,
        html_message,
        settings.DEFAULT_FROM_EMAIL,
        [reservation.email],
    )
    email.content_subtype = "html"  # Set content to HTML
    email.send()

def cancel_confirmed_reservation(request):
    if request.method == 'POST':
        reservation_id = request.POST.get('reservation_id')
        reason = request.POST.get('reason_for_cancelling')
        cancelled_by = request.POST.get('cancelled_by')
        
        try:
            reservation = Reservation.objects.get(id=reservation_id, status='Confirmed')
            reservation.status = 'Cancelled'
            reservation.reason_for_cancelling = reason
            reservation.cancelled_by = cancelled_by
            reservation.cancelled_at = timezone.now()  # Update the cancelled_at datetime field
            reservation.save()
            
            # Send email notification
            send_cancellation_email(reservation, reason)
            
            messages.success(request, f'Confirmed reservation #{reservation_id} has been cancelled successfully. An email has been sent to the guest.')
        except Reservation.DoesNotExist:
            messages.error(request, 'Confirmed reservation not found or already cancelled.')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
    
    # Redirect back to the appropriate page
    if request.META.get('HTTP_REFERER'):
        return redirect(request.META['HTTP_REFERER'])
    return redirect('confirmed_reservations')  # Fallback if no referer

def cancelled_reservations_view(request):
    # Get all cancelled reservations ordered by created_at in descending order
    cancelled_reservations = Reservation.objects.filter(status="Cancelled").order_by('-created_at')

    # Convert 'cancelled_at' to local time for each reservation
    for reservation in cancelled_reservations:
        reservation.created_at = timezone.localtime(reservation.created_at)
        reservation.cancelled_at = timezone.localtime(reservation.cancelled_at)

    # Pagination
    paginator = Paginator(cancelled_reservations, 10)  # 10 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'cancelled_reservations': page_obj
    }
    return render(request, 'administrator/reservation/cancelled_reservations.html', context)

def completed_reservations(request):
    # Get all completed reservations ordered by most recent first
    completed_reservations_list = Reservation.objects.filter(
        status="Completed"
    ).order_by("-created_at")
    
    for reservation in completed_reservations_list:
        reservation.created_at = timezone.localtime(reservation.created_at)
        reservation.completed_at = timezone.localtime(reservation.completed_at)

    # Paginate the results (10 per page)
    paginator = Paginator(completed_reservations_list, 10)
    page_number = request.GET.get('page')
    completed_reservations = paginator.get_page(page_number)
    
    context = {
        'completed_reservations': completed_reservations,
    }
    
    return render(request, 'administrator/reservation/completed_reservations.html', context)

def pending_reservations(request):
    pending_requests = Reservation.objects.filter(status="Pending").order_by("-created_at")
    return render(request, "administrator/reservation/pending_requests.html", {"pending_requests": pending_requests})

@require_POST
def complete_reservation(request):
    reservation_id = request.POST.get('reservation_id')
    
    try:
        reservation = get_object_or_404(Reservation, id=reservation_id)
        
        # Update status and completed_at field
        reservation.status = 'Completed'
        reservation.completed_at = timezone.now()  # Update the completed_at datetime field
        reservation.save()

        messages.success(request, f"Reservation for {reservation.first_name} {reservation.last_name} has been marked as completed.")
    
    except Reservation.DoesNotExist:
        messages.error(request, 'Reservation not found.')
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')

    return redirect('confirmed_reservations')  # Redirect to the appropriate page

def custom_404_view(request, exception):
    return render(request, 'home/404.html', status=404)